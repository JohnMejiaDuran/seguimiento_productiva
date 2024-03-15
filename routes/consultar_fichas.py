from flask import (
    render_template,
    request,
    Blueprint,
    url_for,
    redirect,
    flash,
    session,
    send_file,
)
import pandas as pd
from models.seguimientos import Instructor
from datetime import datetime, timedelta
import xlrd
from io import BytesIO
from xlrd import XLRDError
from models.seguimientos import Aprendiz
from utils.db import db
from flask_login import login_required, current_user
from functools import wraps
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

consultar_ficha = Blueprint("consultar_ficha", __name__)


def admin_required(fn):
    @wraps(fn)
    def decorated_view(*args, **kwargs):
        if current_user.is_authenticated and any(
            role.name == "Administrador" for role in current_user.roles
        ):
            return fn(*args, **kwargs)
        else:
            flash("No tienes acceso a esta página.")
            return redirect(url_for("pagina_inicio.index"))

    return decorated_view


@consultar_ficha.route("/descargar_excel", methods=["GET"])
@login_required
@admin_required
def descargar_excel():
    # Obtener los datos de la sesión
    datos_excel = session.get("datos_excel")

    # Comprobar si hay datos en la sesión
    if datos_excel:
        # Crear un DataFrame con los datos de la sesión
        df = pd.DataFrame(datos_excel)

        # Nombre del archivo Excel de salida
        nombre_archivo_excel = "datos_descarga.xlsx"

        # Guardar el DataFrame en un archivo Excel
        df.to_excel(nombre_archivo_excel, index=False)

        # Enviar el archivo Excel como una respuesta al cliente
        return send_file(nombre_archivo_excel, as_attachment=True)

    else:
        # Si no hay datos en la sesión, redireccionar a la página de origen
        flash("No se encontraron datos para descargar", "error")
        return redirect(url_for("consultar_ficha.consultarficha"))


@consultar_ficha.route("/consultar_fichas", methods=["GET", "POST"])
@login_required
@admin_required
def consultarficha():
    titulo = "Consultar por fichas"
    logo = "/static/icons/user-icon.png"
    title = "Consultar fichas"

    xls = session.pop("xls_data", True)
    not_data = session.pop("not_data", True)
    return render_template(
        "consultar_fichas.html",
        titulo=titulo,
        logo=logo,
        title=title,
        xls=xls,
        not_data=not_data,
    )


# debo hacer otra ruta donde este solo la tabla y esta extraiga los datos  de la ruta llamada table que deberia llamarse validacion datos para observarlos


@consultar_ficha.route("/table", methods=["GET", "POST"])
@login_required
@admin_required
def table():
    rol = ""
    instructores = Instructor.query.all()
    if request.method == "POST":
        # Verifica si se ha enviado un archivo en la solicitud.

        if "archivo" in request.files:
            archivo = request.files["archivo"]
            try:
                if not archivo.filename.endswith((".xls")):
                    xls = False
                    flash("Error al procesar el archivo .xls:", "error")
                    session["xls_data"] = xls
                    return redirect(url_for("consultar_ficha.consultarficha", xls=xls))

                contenido_archivo = archivo.read()
                memoria_archivo = BytesIO(contenido_archivo)
                workbook_xls = xlrd.open_workbook(file_contents=memoria_archivo.read())

                sheet = workbook_xls.sheet_by_index(0)
                ficha = sheet.cell_value(2, 2)
                ficha_sin_decimal = str(int(ficha))
                modalidad = sheet.cell_value(9, 2)
                programa = sheet.cell_value(5, 2)
                fecha_inicio = sheet.cell_value(7, 2)
                fecha_excel = sheet.cell_value(8, 2)
                regional = sheet.cell_value(10, 2)
                centro = sheet.cell_value(11, 2)
                centro_codigo_primeras_4 = centro[:4]
                codigo_centro = int(centro_codigo_primeras_4)

                if isinstance(fecha_inicio, float):
                    fecha_python_1 = xlrd.xldate_as_datetime(fecha_inicio, 0)
                    fecha_sin_hora_inicio = fecha_python_1.date()

                if isinstance(fecha_excel, float):
                    fecha_python = xlrd.xldate_as_datetime(
                        fecha_excel, 0
                    )  # El segundo argumento es el modo de fecha en Excel
                    fecha_sin_hora = fecha_python.date()
                    fecha_actual = datetime.now().date()

                    ficha_en_vigencia = True
                    if fecha_sin_hora + timedelta(days=548) < fecha_actual:
                        ficha_en_vigencia = False

                    else:
                        print(
                            "No es un valor de fecha en Excel en la celda especificada."
                        )
            except (ValueError, KeyError) as e:
                flash(f"Error en los datos del archivo Excel: {str(e)}", "error")

            if archivo:
                try:
                    df = pd.read_excel(archivo, skiprows=12)

                    #########################################################################
                    # APRENDICES CON ETAPA ELECTIVA APROBADA

                    en_formacion = df[
                        (df["Estado"] == "EN FORMACION")
                        & (
                            df["Competencia"]
                            != "2 - RESULTADOS DE APRENDIZAJE ETAPA PRÁCTICA"
                        )
                    ]  # primero lista los aprendices con estado en formacion y no tiene en cuenta el resultado de la etapa practica

                    conteo_por_persona_aprobada = (
                        en_formacion[en_formacion["Juicio de Evaluación"] == "APROBADO"]
                        .groupby(
                            ["Nombre", "Número de Documento", "Estado", "Apellidos"]
                        )
                        .size()
                    )  # el aprendiz debe tener los jucios de evaluacion aprobados y los agrupa por nombre, documento, estado y apellidos

                    documentos = (
                        en_formacion.groupby("Nombre")["Número de Documento"].size() - 1
                    )  # Cuenta cuantos juicios tiene cada aprendiz mediante la cedula y le resta 1 que es la practica

                    aprendices_aprobados = conteo_por_persona_aprobada[
                        conteo_por_persona_aprobada.eq(documentos)
                    ].reset_index()  # conteo por persona aprobada debe ser igual al conteo de documentos -1 ya que no contamos el resultado de la etapa practica
                    aprendices_aprobados["Ficha"] = ficha_sin_decimal.strip()

                    # Cargar los números de documento de la base de datos SQLAlchemy
                    numeros_documento_bd = db.session.query(
                        Aprendiz.documento
                    ).all()  # Suponiendo que 'session' es tu sesión SQLAlchemy y 'Aprendiz' es tu tabla

                    numeros_documento_bd_str = [
                        str(num[0]) for num in numeros_documento_bd
                    ]
                    numeros_ficha_bd = db.session.query(Aprendiz.ficha_id).all()
                    numeros_ficha_bd_str = [str(num[0]) for num in numeros_ficha_bd]
                    
                    # Filtrar los aprendices basados en si sus números de documento ya están en la base de datos SQLAlchemy
                    aprendices_no_en_bd = aprendices_aprobados[
                            ~aprendices_aprobados["Ficha"]
                            .astype(str)
                            .isin(numeros_ficha_bd_str)
                        ]

                    print("APRENDICES DOCUMENTO BD STR")
                    print(numeros_documento_bd_str)
                    print("APRENDICES DOCUMENTO BD")
                    print(numeros_documento_bd)
                    print("APRENDICES APROBADOS")
                    print(aprendices_aprobados)
                    print(" FICHA   EN BASE DE DATOS")
                    print(type(numeros_ficha_bd_str))
                    print(numeros_ficha_bd)
                    print("NO EN BD")
                    print(aprendices_no_en_bd)
                    #########################################################################
                    # APRENDICES CON JUICIOS PENDIENTES

                    competencias_por_evaluar = df[
                        (df["Juicio de Evaluación"] == "POR EVALUAR")
                        & (df["Estado"] == "EN FORMACION")
                        & (
                            df["Competencia"]
                            != "2 - RESULTADOS DE APRENDIZAJE ETAPA PRACTICA"
                        )
                    ]

                    practicasporevaluar = df[
                        (
                            df["Competencia"]
                            == "2 - RESULTADOS DE APRENDIZAJE ETAPA PRACTICA"
                        )
                        & (df["Juicio de Evaluación"] == "POR EVALUAR")
                    ]

                    evaluar = competencias_por_evaluar[
                        ~(
                            competencias_por_evaluar["Número de Documento"].isin(
                                practicasporevaluar
                            )
                        )
                    ]

                    #########################################################################
                    # APRENDICES CON NOVEDADES
                    #
                    aprendices_novedades = df[df["Estado"] != "EN FORMACION"]

                    # Luego puedes aplicar drop_duplicates
                    novedades = aprendices_novedades.drop_duplicates(
                        subset=("Número de Documento")
                    ).reset_index()

                    # Extracción de datos, ficha, fechas y validar si la ficha aun es vigente

                    # Crear archivos Excel para aprendices con novedades y juicios pendientes
                    nuevo_libro = Workbook()
                    hoja_nueva = nuevo_libro.active

                    # Establecer los valores en celdas específicas
                    hoja_nueva["A1"] = "Aprendices con Juicios Pendientes"
                    hoja_nueva.merge_cells("A1:H1")

                    hoja_nueva["A2"] = "Fecha del Reporte:"
                    hoja_nueva.merge_cells("A2:B2")
                    hoja_nueva["C2"] = fecha_actual

                    hoja_nueva["A3"] = "Ficha de Caracterización:"
                    hoja_nueva.merge_cells("A3:B3")
                    hoja_nueva["C3"] = ficha_sin_decimal
                    hoja_nueva.merge_cells("C3:F3")

                    hoja_nueva["A4"] = "Programa:"
                    hoja_nueva.merge_cells("A4:B4")
                    hoja_nueva["C4"] = programa
                    hoja_nueva.merge_cells("C4:F4")

                    hoja_nueva["A5"] = "Modalidad de Formación:"
                    hoja_nueva.merge_cells("A5:B5")
                    hoja_nueva["C5"] = modalidad
                    hoja_nueva.merge_cells("C5:F5")

                    hoja_nueva["A6"] = "Fecha Inicio:"
                    hoja_nueva.merge_cells("A6:B6")
                    hoja_nueva["C6"] = fecha_sin_hora_inicio

                    hoja_nueva["A7"] = "Fecha Fin:"
                    hoja_nueva.merge_cells("A7:B7")
                    hoja_nueva["C7"] = fecha_sin_hora

                    hoja_nueva["A8"] = "Regional:"
                    hoja_nueva.merge_cells("A8:B8")
                    hoja_nueva["C8"] = regional
                    hoja_nueva.merge_cells("C8:F8")

                    hoja_nueva["A9"] = "Centro:"
                    hoja_nueva.merge_cells("A9:B9")
                    hoja_nueva["C9"] = centro
                    hoja_nueva.merge_cells("C9:F9")

                    hoja_nueva["A10"] = "Tipo de Documento"

                    hoja_nueva["B10"] = "Número de Documento"

                    hoja_nueva["C10"] = "Nombre"

                    hoja_nueva["D10"] = "Apellidos"

                    hoja_nueva["E10"] = "Estado"

                    hoja_nueva["F10"] = "Competencia"

                    hoja_nueva["G10"] = "Resultado de Aprendizaje"

                    hoja_nueva["H10"] = "Juicio de Evaluación"

                    fila_actual = 11
                    for index, row in evaluar.iterrows():
                        hoja_nueva["A{}".format(fila_actual)] = row["Tipo de Documento"]
                        hoja_nueva["B{}".format(fila_actual)] = row[
                            "Número de Documento"
                        ]
                        hoja_nueva["C{}".format(fila_actual)] = row["Nombre"]
                        hoja_nueva["D{}".format(fila_actual)] = row["Apellidos"]
                        hoja_nueva["E{}".format(fila_actual)] = row["Estado"]
                        hoja_nueva["F{}".format(fila_actual)] = row["Competencia"]
                        hoja_nueva["G{}".format(fila_actual)] = row[
                            "Resultado de Aprendizaje"
                        ]
                        hoja_nueva["H{}".format(fila_actual)] = row[
                            "Juicio de Evaluación"
                        ]
                        fila_actual += 1
                        # Incrementar el contador de fila
                    for col in range(1, hoja_nueva.max_column + 1):
                        max_length = 0
                        columna_letra = get_column_letter(col)
                        for row in range(
                            10, hoja_nueva.max_row + 1
                        ):  # Empezar desde la fila 10
                            cell = hoja_nueva[columna_letra + str(row)]
                            if (
                                cell.value is not None
                            ):  # Verificar si el valor de la celda no es nulo
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        adjusted_width = (max_length + 2) * 1.2
                        hoja_nueva.column_dimensions[columna_letra].width = (
                            adjusted_width
                        )
                    nuevo_libro.save("juicios_pendientes.xlsx")

                    nuevo_libro2 = Workbook()
                    hoja_nueva2 = nuevo_libro2.active

                    hoja_nueva2["A1"] = "Aprendices con Juicios Pendientes"
                    hoja_nueva2.merge_cells("A1:H1")

                    hoja_nueva2["A2"] = "Fecha del Reporte:"
                    hoja_nueva2.merge_cells("A2:B2")
                    hoja_nueva2["C2"] = fecha_actual

                    hoja_nueva2["A3"] = "Ficha de Caracterización:"
                    hoja_nueva2.merge_cells("A3:B3")
                    hoja_nueva2["C3"] = ficha_sin_decimal
                    hoja_nueva2.merge_cells("C3:F3")

                    hoja_nueva2["A4"] = "Programa:"
                    hoja_nueva2.merge_cells("A4:B4")
                    hoja_nueva2["C4"] = programa
                    hoja_nueva2.merge_cells("C4:F4")

                    hoja_nueva2["A5"] = "Modalidad de Formación:"
                    hoja_nueva2.merge_cells("A5:B5")
                    hoja_nueva2["C5"] = modalidad
                    hoja_nueva2.merge_cells("C5:F5")

                    hoja_nueva2["A6"] = "Fecha Inicio:"
                    hoja_nueva2.merge_cells("A6:B6")
                    hoja_nueva2["C6"] = fecha_sin_hora_inicio

                    hoja_nueva2["A7"] = "Fecha Fin:"
                    hoja_nueva2.merge_cells("A7:B7")
                    hoja_nueva2["C7"] = fecha_sin_hora

                    hoja_nueva2["A8"] = "Regional:"
                    hoja_nueva2.merge_cells("A8:B8")
                    hoja_nueva2["C8"] = regional
                    hoja_nueva2.merge_cells("C8:F8")

                    hoja_nueva2["A9"] = "Centro:"
                    hoja_nueva2.merge_cells("A9:B9")
                    hoja_nueva2["C9"] = centro
                    hoja_nueva2.merge_cells("C9:F9")

                    hoja_nueva2["A10"] = "Tipo de Documento"

                    hoja_nueva2["B10"] = "Número de Documento"

                    hoja_nueva2["C10"] = "Nombre"

                    hoja_nueva2["D10"] = "Apellidos"

                    hoja_nueva2["E10"] = "Estado"

                    fila_actual = 11
                    for index, row in novedades.iterrows():
                        hoja_nueva2["A{}".format(fila_actual)] = row[
                            "Tipo de Documento"
                        ]
                        hoja_nueva2["B{}".format(fila_actual)] = row[
                            "Número de Documento"
                        ]
                        hoja_nueva2["C{}".format(fila_actual)] = row["Nombre"]
                        hoja_nueva2["D{}".format(fila_actual)] = row["Apellidos"]
                        hoja_nueva2["E{}".format(fila_actual)] = row["Estado"]

                        fila_actual += 1
                    nuevo_libro2.save("novedades.xlsx")
                    return render_template(
                        "table.html",
                        fecha_actual=fecha_actual,
                        fecha_sin_hora_inicio=fecha_sin_hora_inicio,
                        fecha_sin_hora=fecha_sin_hora,
                        aprendices_no_en_bd=aprendices_no_en_bd,
                        evaluar=evaluar,
                        novedades=novedades,
                        codigo_centro=codigo_centro,
                        rol=rol,
                        instructores=instructores,
                        ficha_sin_decimal=ficha_sin_decimal,
                        programa=programa,
                        ficha_en_vigencia=ficha_en_vigencia,
                    )
                except (ValueError, KeyError, TypeError):
                    not_data = False
                    flash("Error en los datos del archivo Excel", "error")
                    session["not_data"] = not_data
                    return redirect(
                        url_for("consultar_ficha.consultarficha", not_data=not_data)
                    )

    else:
        return redirect(url_for("consultar_ficha.consultarficha"))


@consultar_ficha.route("/descargar_datos")
@login_required
def descargar_datos():
    # Ruta al archivo juicios_pendientes.xlsx
    ruta_archivo = "juicios_pendientes.xlsx"
    # Utiliza Flask's send_file para enviar el archivo al cliente para su descarga
    return send_file(ruta_archivo, as_attachment=True)


@consultar_ficha.route("/descargar_datos_novedades")
@login_required
def descargar_datos_novedades():
    # Ruta al archivo novedades.xlsx
    ruta_archivo = "novedades.xlsx"
    # Utiliza Flask's send_file para enviar el archivo al cliente para su descarga
    return send_file(ruta_archivo, as_attachment=True)
